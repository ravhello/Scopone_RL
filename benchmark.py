#!/usr/bin/env python3
"""
Checkpoint Benchmark Script for Scopone AI

This script compares different checkpoints of the Scopone AI model by having them
play against each other to benchmark their performance.

Usage:
  python benchmark.py --checkpoint_dir checkpoints/ --games 1000
  python benchmark.py --checkpoints checkpoints/model_team0_ep5000.pth checkpoints/model_team0_ep10000.pth
"""

import torch
import os
import math
import argparse
import time
import re
from tqdm import tqdm
import itertools
import glob
import pandas as pd
from typing import Optional, List
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import openpyxl

# Import the required components from the existing code
from environment import ScoponeEnvMA
from algorithms.is_mcts import run_is_mcts
from models.action_conditioned import ActionConditionedActor, CentralValueNet
from utils.compile import maybe_compile_module
from utils.torch_load import safe_torch_load
from evaluation.eval import evaluate_pair_actors_parallel, evaluate_pair_actors
from selfplay.league import League


def _ensure_default_eval_env() -> None:
    """Align evaluation defaults with main.py so benchmark uses identical settings."""
    defaults = {
        'SCOPONE_TORCH_COMPILE': '0',
        'SCOPONE_TORCH_COMPILE_MODE': 'reduce-overhead',
        'SCOPONE_TORCH_COMPILE_BACKEND': 'inductor',
        'SCOPONE_INDUCTOR_AUTOTUNE': '1',
        'SCOPONE_EVAL_GAMES': '1000',
        'SCOPONE_EVAL_K_HISTORY': '39',
        'SCOPONE_EVAL_USE_MCTS': '0',
        'SCOPONE_EVAL_MCTS_C_PUCT': '1.0',
        'SCOPONE_EVAL_MCTS_ROOT_TEMP': '0.0',
        'SCOPONE_EVAL_MCTS_PRIOR_SMOOTH_EPS': '0.0',
        'SCOPONE_EVAL_MCTS_DIRICHLET_ALPHA': '0.25',
        'SCOPONE_EVAL_MCTS_DIRICHLET_EPS': '0.25',
        'SCOPONE_EVAL_MCTS_SIMS': '4',
        'SCOPONE_EVAL_MCTS_DETS_PRIOR': '2',
        'SCOPONE_EVAL_BELIEF_PARTICLES': '0',
        'SCOPONE_EVAL_BELIEF_ESS_FRAC': '0.5',
        'SCOPONE_EVAL_MCTS_SCALING': '1',
        'SCOPONE_EVAL_MCTS_PROGRESS_START': '0.25',
        'SCOPONE_EVAL_MCTS_PROGRESS_FULL': '0.75',
        'SCOPONE_EVAL_MCTS_MIN_SIMS': '0',
        'SCOPONE_EVAL_MCTS_TRAIN_FACTOR': '1.0',
        'SCOPONE_EVAL_POOL_TIMEOUT_S': '0',
        'SCOPONE_EVAL_MCTS_EXACT_MAX_MOVES': '12',
        'SCOPONE_EVAL_MCTS_EXACT_ONLY': '1',
        'SCOPONE_EVAL_MCTS_EXACT_COVER_FRAC': '70',
        'SCOPONE_EVAL_MCTS_DETS_EXACT': '4',
        'SCOPONE_EVAL_WORKERS': str(max(1, (os.cpu_count() or 1))),
        'SCOPONE_ELO_DIFF_SCALE': '6.0',
        'TQDM_DISABLE': '0',
        'SCOPONE_WORKER_THREADS': '1',
    }
    for key, value in defaults.items():
        os.environ.setdefault(key, value)
    # Force evaluation to run on CPU for stability (match main.py defaults)
    os.environ['SCOPONE_DEVICE'] = 'cpu'
    os.environ['SCOPONE_TRAIN_DEVICE'] = 'cpu'
    os.environ['ENV_DEVICE'] = 'cpu'
    # Hide CUDA devices unless explicitly re-enabled by the user
    if os.environ.get('CUDA_VISIBLE_DEVICES', None) not in ('', '-1'):
        os.environ['CUDA_VISIBLE_DEVICES'] = ''


_ensure_default_eval_env()


def _build_eval_mcts_cfg() -> Optional[dict]:
    """Construct an evaluation MCTS config mirroring main.py defaults."""
    use_mcts = str(os.environ.get('SCOPONE_EVAL_USE_MCTS', '1')).strip().lower() in ['1', 'true', 'yes', 'on']
    if not use_mcts:
        return None
    try:
        sims = int(os.environ.get('SCOPONE_EVAL_MCTS_SIMS', '4'))
    except ValueError:
        sims = 4
    try:
        dets = int(os.environ.get('SCOPONE_EVAL_MCTS_DETS_PRIOR', '2'))
    except ValueError:
        dets = 2
    cfg = {
        'sims': max(0, sims),
        'dets': max(1, dets),
        'c_puct': float(os.environ.get('SCOPONE_EVAL_MCTS_C_PUCT', '1.0')),
        'root_temp': float(os.environ.get('SCOPONE_EVAL_MCTS_ROOT_TEMP', '0.0')),
        'prior_smooth_eps': float(os.environ.get('SCOPONE_EVAL_MCTS_PRIOR_SMOOTH_EPS', '0.0')),
        'root_dirichlet_alpha': float(os.environ.get('SCOPONE_EVAL_MCTS_DIRICHLET_ALPHA', '0.25')),
        'root_dirichlet_eps': float(os.environ.get('SCOPONE_EVAL_MCTS_DIRICHLET_EPS', '0.25')),
        'robust_child': True,
        'progress_start': float(os.environ.get('SCOPONE_EVAL_MCTS_PROGRESS_START', '0.25')),
        'progress_full': float(os.environ.get('SCOPONE_EVAL_MCTS_PROGRESS_FULL', '0.75')),
        'min_sims': int(os.environ.get('SCOPONE_EVAL_MCTS_MIN_SIMS', '0')),
        'train_factor': float(os.environ.get('SCOPONE_EVAL_MCTS_TRAIN_FACTOR', '1.0')),
    }
    return cfg
 

# Device selection with overrides
# Force CPU device regardless of environment or CUDA availability
device = torch.device('cpu')
#print(f"Using device: {device}")

# Regole di default per l'ambiente (modalità standard senza varianti)
# Nota: la variante "asso_piglia_tutto" è disattivata e quindi ignorata in questo script di benchmark.
DEFAULT_RULES = {
    'start_with_4_on_table': False,
    'asso_piglia_tutto': False,
    'scopa_on_asso_piglia_tutto': False,
    'asso_piglia_tutto_posabile': False,
    'asso_piglia_tutto_posabile_only_empty': False,
    'scopa_on_last_capture': False,
    're_bello': False,
    'napola': False,
    'napola_scoring': 'fixed3',
    'max_consecutive_scope': None,
    'last_cards_to_dealer': True,
}

def load_actor_critic(ckpt_path: str):
    actor = ActionConditionedActor()
    critic = CentralValueNet()
    # Enable compiled versions when requested (works on CPU with inductor)
    actor = maybe_compile_module(actor, name='ActionConditionedActor[benchmark]')
    critic = maybe_compile_module(critic, name='CentralValueNet[benchmark]')
    ckpt = safe_torch_load(ckpt_path, map_location=device)
    # se si usa algorithms/ppo_ac save
    if 'actor' in ckpt and 'critic' in ckpt:
        actor.load_state_dict(ckpt['actor'])
        critic.load_state_dict(ckpt['critic'])
    else:
        raise RuntimeError('No weights in checkpoint')
    return actor, critic

def play_game(actor1, actor2, starting_player=0, use_mcts=False, sims=128, dets=16):
    """
    Play a single game between two Team 0 agents.
    
    Args:
        agent1: First agent
        agent2: Second agent
        starting_player: Which player starts (0-3)
    
    Returns:
        winner: 0 if agent1 won, 1 if agent2 won, -1 if draw
        team_scores: Scores for each agent [agent1_score, agent2_score]
        game_length: Number of moves in the game
    """
    env = ScoponeEnvMA(rules=DEFAULT_RULES)
    env.current_player = starting_player
    
    done = False
    info = {}
    
    # Track which actor controls which seats
    actor1_positions = [0, 2]  # Team 0 positions
    actor2_positions = [1, 3]  # Team 1 positions
    
    # Game loop
    while not done:
        current_player = env.current_player
        
        # Get valid actions
        valid_actions = env.get_valid_actions()
        
        if len(valid_actions) == 0:
            print("\n[ERROR] No valid actions available!")
            break
        
        # Get observation for current player
        obs = env._get_observation(current_player)
        
        # Select action (with optional IS-MCTS if available)
        if use_mcts:
            # Use the corresponding loaded actor; critic can be a lightweight instance for value
            actor = actor1 if current_player in actor1_positions else actor2
            critic = maybe_compile_module(CentralValueNet(), name='CentralValueNet[bench_mcts]')
            def policy_fn(o, leg):
                o_t = torch.tensor(o, dtype=torch.float32, device=device)
                if len(leg) > 0 and torch.is_tensor(leg[0]):
                    leg_t = torch.stack(leg).to(device=device, dtype=torch.float32)
                else:
                    leg_t = torch.stack([
                        x if torch.is_tensor(x) else torch.tensor(x, dtype=torch.float32, device=device)
                    for x in leg], dim=0)
                with torch.no_grad():
                    # seat/team vec aligned to current env seat
                    s = torch.zeros(6, dtype=torch.float32, device=device)
                    cp = env.current_player
                    s[cp] = 1.0
                    s[4] = 1.0 if cp in [0, 2] else 0.0
                    s[5] = 1.0 if cp in [1, 3] else 0.0
                    logits = actor(o_t.unsqueeze(0), leg_t, s.unsqueeze(0))
                probs = torch.softmax(logits, dim=0)
                return probs
            def value_fn(o, _env=None):
                o_t = o.clone().detach().to(device=device, dtype=torch.float32) if torch.is_tensor(o) else torch.tensor(o, dtype=torch.float32, device=device)
                # build seat vector if env available
                if _env is not None:
                    cp = _env.current_player
                    s = torch.zeros(6, dtype=torch.float32, device=device)
                    s[cp] = 1.0
                    s[4] = 1.0 if cp in [0, 2] else 0.0
                    s[5] = 1.0 if cp in [1, 3] else 0.0
                else:
                    s = torch.zeros(6, dtype=torch.float32, device=device)
                with torch.no_grad():
                    return critic(o_t.unsqueeze(0), s.unsqueeze(0)).item()
            # Belief sampler neurale: determinizza le mani avversarie dai margini del BeliefNet
            def belief_sampler_neural(_env):
                cp = _env.current_player
                obs_cur = _env._get_observation(cp)
                o_cpu = obs_cur if torch.is_tensor(obs_cur) else torch.as_tensor(obs_cur, dtype=torch.float32)
                if torch.is_tensor(o_cpu):
                    o_cpu = o_cpu.detach().to('cpu', dtype=torch.float32)
                # seat/team vec: 6-dim
                s_cpu = torch.zeros(6, dtype=torch.float32)
                s_cpu[cp] = 1.0
                s_cpu[4] = 1.0 if cp in [0, 2] else 0.0
                s_cpu[5] = 1.0 if cp in [1, 3] else 0.0
                if device.type == 'cuda':
                    o_t = o_cpu.pin_memory().unsqueeze(0).to(device=device, non_blocking=True)
                    s_t = s_cpu.pin_memory().unsqueeze(0).to(device=device, non_blocking=True)
                else:
                    o_t = o_cpu.unsqueeze(0).to(device=device)
                    s_t = s_cpu.unsqueeze(0).to(device=device)
                with torch.no_grad():
                    state_feat = actor.state_enc(o_t, s_t)
                    logits = actor.belief_net(state_feat)
                    hand_table = o_t[:, :83]
                    hand_mask = hand_table[:, :40] > 0.5
                    table_mask = hand_table[:, 43:83] > 0.5
                    captured = o_t[:, 83:165]
                    cap0_mask = captured[:, :40] > 0.5
                    cap1_mask = captured[:, 40:80] > 0.5
                    visible_mask = (hand_mask | table_mask | cap0_mask | cap1_mask)
                    probs_flat = actor.belief_net.probs(logits, visible_mask)
                probs = probs_flat.view(3, 40).detach().cpu().numpy()
                vis = visible_mask.squeeze(0).detach().cpu().numpy().astype(bool)
                unknown_ids = [cid for cid in range(40) if not vis[cid]]
                others = [(cp + 1) % 4, (cp + 2) % 4, (cp + 3) % 4]
                det = {pid: [] for pid in others}
                counts = {pid: len(_env.game_state['hands'][pid]) for pid in others}
                caps = [int(counts.get(pid, 0)) for pid in others]
                n = len(unknown_ids)
                if sum(caps) != n:
                    caps[2] = max(0, n - caps[0] - caps[1])
                log_prob = 0.0
                for cid in unknown_ids:
                    pc = probs[:, cid]
                    s = pc.sum()
                    ps = pc / (s if s > 0 else 1e-9)
                    j = int(torch.argmax(torch.tensor(ps)).item())
                    if caps[j] > 0:
                        det[others[j]].append(cid)
                        caps[j] -= 1
                        log_prob += math.log(max(1e-12, float(ps[j])))
                return {'assignments': det, 'logp': log_prob}
            action = run_is_mcts(env, policy_fn, value_fn, num_simulations=sims, c_puct=1.0, belief=None, num_determinization=dets,
                                    belief_sampler=belief_sampler_neural)
        else:
            # Greedy selection by actor scoring
            with torch.no_grad():
                o_t = obs.clone().detach().to(device=device, dtype=torch.float32) if torch.is_tensor(obs) else torch.tensor(obs, dtype=torch.float32, device=device)
                leg_t = torch.stack([
                    x if torch.is_tensor(x) else torch.tensor(x, dtype=torch.float32, device=device)
                for x in valid_actions], dim=0)
                s = torch.zeros(6, dtype=torch.float32, device=device)
                s[current_player] = 1.0
                s[4] = 1.0 if current_player in [0, 2] else 0.0
                s[5] = 1.0 if current_player in [1, 3] else 0.0
                actor = actor1 if current_player in actor1_positions else actor2
                logits = actor(o_t.unsqueeze(0), leg_t, s.unsqueeze(0))
                idx = torch.argmax(logits).item()
            action = valid_actions[idx]

        # Take step in environment
        next_obs, reward, done, info = env.step(action)

    # Extract final score information from team_rewards
    agent1_score = 0.0
    agent2_score = 0.0
    if "team_rewards" in info:
        team_rewards = info["team_rewards"]
        agent1_score = team_rewards[0]  # Team 0 score for agent1
        agent2_score = team_rewards[1]  # Team 1 score for agent2
    
    # Determine winner based on team scores
    winner = 0 if agent1_score > agent2_score else 1 if agent2_score > agent1_score else -1
    
    return winner, [agent1_score, agent2_score], len(env.game_state["history"])
def main_cli():
    import argparse
    parser = argparse.ArgumentParser(description='Benchmark Scopone with optional IS-MCTS')
    parser.add_argument('--mcts', action='store_true', help='Use IS-MCTS booster')
    parser.add_argument('--sims', type=int, default=128, help='Number of MCTS simulations')
    parser.add_argument('--dets', type=int, default=16, help='Number of belief determinisations per search')
    parser.add_argument('--compact', action='store_true', help='Use compact observation')
    parser.add_argument('--k-history', type=int, default=39, help='Recent moves for compact observation')
    parser.add_argument('--ckpt', type=str, default='', help='Checkpoint path for actor/critic (optional)')
    parser.add_argument('--games', type=int, default=1000, help='Number of games to play')
    args = parser.parse_args()
    # Placeholder for agent loading and running a quick game
    actor, critic = load_actor_critic(args.ckpt) if args.ckpt else (maybe_compile_module(ActionConditionedActor(), name='ActionConditionedActor[benchmark]'),
                                                                     maybe_compile_module(CentralValueNet(), name='CentralValueNet[benchmark]'))
    for g in range(args.games):
        env = ScoponeEnvMA(k_history=args.k_history)
        done = False
        while not done:
            obs = env._get_observation(env.current_player)
            legals = env.get_valid_actions()
            if len(legals) == 0:
                break
            if args.mcts:
                def policy_fn(o, leg):
                    o_t = o.clone().detach().to(device=device, dtype=torch.float32) if torch.is_tensor(o) else torch.tensor(o, dtype=torch.float32, device=device)
                    leg_t = torch.stack([
                        x if torch.is_tensor(x) else torch.tensor(x, dtype=torch.float32, device=device)
                    for x in legals], dim=0)
                    with torch.no_grad():
                        logits = actor(o_t.unsqueeze(0), leg_t)
                    return torch.softmax(logits, dim=0)
                def value_fn(o):
                    o_t = torch.tensor(o, dtype=torch.float32, device=device)
                    # derive a simple seat vector from env state is not available here; use zeros
                    s = torch.zeros(6, dtype=torch.float32, device=device)
                    with torch.no_grad():
                        return critic(o_t.unsqueeze(0), s.unsqueeze(0)).item()
                def belief_sampler_neural(_env):
                    cp = _env.current_player
                    obs_cur = _env._get_observation(cp)
                    o_cpu = obs_cur if torch.is_tensor(obs_cur) else torch.as_tensor(obs_cur, dtype=torch.float32)
                    if torch.is_tensor(o_cpu):
                        o_cpu = o_cpu.detach().to('cpu', dtype=torch.float32)
                    s_cpu = torch.zeros(6, dtype=torch.float32)
                    s_cpu[cp] = 1.0
                    s_cpu[4] = 1.0 if cp in [0, 2] else 0.0
                    s_cpu[5] = 1.0 if cp in [1, 3] else 0.0
                    if device.type == 'cuda':
                        o_t = o_cpu.pin_memory().unsqueeze(0).to(device=device, non_blocking=True)
                        s_t = s_cpu.pin_memory().unsqueeze(0).to(device=device, non_blocking=True)
                    else:
                        o_t = o_cpu.unsqueeze(0).to(device=device)
                        s_t = s_cpu.unsqueeze(0).to(device=device)
                    with torch.no_grad():
                        state_feat = actor.state_enc(o_t, s_t)
                        logits = actor.belief_net(state_feat)
                        hand_table = o_t[:, :83]
                        hand_mask = hand_table[:, :40] > 0.5
                        table_mask = hand_table[:, 43:83] > 0.5
                        captured = o_t[:, 83:165]
                        cap0_mask = captured[:, :40] > 0.5
                        cap1_mask = captured[:, 40:80] > 0.5
                        visible_mask = (hand_mask | table_mask | cap0_mask | cap1_mask)
                        probs_flat = actor.belief_net.probs(logits, visible_mask)
                    probs = probs_flat.view(3, 40).detach().cpu().numpy()
                    vis = visible_mask.squeeze(0).detach().cpu().numpy().astype(bool)
                    unknown_ids = [cid for cid in range(40) if not vis[cid]]
                    others = [(cp + 1) % 4, (cp + 2) % 4, (cp + 3) % 4]
                    det = {pid: [] for pid in others}
                    counts = {pid: len(_env.game_state['hands'][pid]) for pid in others}
                    caps = [int(counts.get(pid, 0)) for pid in others]
                    n = len(unknown_ids)
                    if sum(caps) != n:
                        caps[2] = max(0, n - caps[0] - caps[1])
                    log_prob = 0.0
                    for cid in unknown_ids:
                        pc = probs[:, cid]
                        s = pc.sum()
                        ps = pc / (s if s > 0 else 1e-9)
                        j = int(torch.argmax(torch.tensor(ps)).item())
                        if caps[j] > 0:
                            det[others[j]].append(cid)
                            caps[j] -= 1
                            log_prob += math.log(max(1e-12, float(ps[j])))
                    return {'assignments': det, 'logp': log_prob}
                action = run_is_mcts(env, policy_fn, value_fn, num_simulations=args.sims, c_puct=1.0, belief=None, num_determinization=args.dets,
                                      belief_sampler=belief_sampler_neural)
            else:
                # pick best by actor scoring
                with torch.no_grad():
                    o_t = obs.clone().detach().to(device=device, dtype=torch.float32) if torch.is_tensor(obs) else torch.tensor(obs, dtype=torch.float32, device=device)
                    leg_t = torch.stack([
                        x if torch.is_tensor(x) else torch.tensor(x, dtype=torch.float32, device=device)
                    for x in legals], dim=0)
                    logits = actor(o_t.unsqueeze(0), leg_t)
                    idx = torch.argmax(logits).item()
                action = legals[idx]
            _, _, done, _ = env.step(action)
    print('Benchmark completed.')

def find_checkpoints(checkpoint_dir, pattern="*.pth"):
    """Find Team 0 checkpoint files in the specified directory."""
    if os.path.isfile(checkpoint_dir):
        return [os.path.normpath(checkpoint_dir)]
    
    checkpoint_pattern = os.path.join(checkpoint_dir, pattern)
    checkpoint_files = glob.glob(checkpoint_pattern)
    
    # Sort checkpoints by episode number when possible
    try:
        checkpoint_files.sort(key=lambda x: int(x.split('_ep')[1].split('.pth')[0]) 
                             if '_ep' in x else float('inf'))
    except Exception as e:
        raise RuntimeError('Failed to sort checkpoint files by episode number') from e
    checkpoint_files = [os.path.normpath(p) for p in checkpoint_files]
    
    return checkpoint_files

def extract_episode_number(checkpoint_path):
    """Extract episode number from checkpoint filename."""
    # Try to find episode number using regex
    match = re.search(r'_ep(\d+)', os.path.basename(checkpoint_path))
    if match:
        return int(match.group(1))
    
    # If not found, try to extract any number from the filename
    match = re.search(r'(\d+)', os.path.basename(checkpoint_path))
    if match:
        return int(match.group(1))
    
    # If still no number, return a very large number to place at the end
    return float('inf')


def checkpoint_display_name(checkpoint_path: str) -> str:
    """Return a readable name for a checkpoint based on its episode number."""
    episode = extract_episode_number(checkpoint_path)
    return f"ep{episode}" if episode != float('inf') else os.path.basename(checkpoint_path).replace(".pth", "")

def canonicalize_league_paths(league: League) -> None:
    """Normalize league history/Elo paths to the local OS format."""
    new_history: List[str] = []
    for path in getattr(league, 'history', []):
        canon = os.path.normpath(path)
        if canon not in new_history:
            new_history.append(canon)
    new_elo = {}
    for path, value in getattr(league, 'elo', {}).items():
        canon = os.path.normpath(path)
        new_elo[canon] = value
    if new_history != league.history or set(new_elo.keys()) != set(league.elo.keys()):
        league.history = new_history
        league.elo = new_elo
        try:
            league._save()
        except Exception:
            pass

def evaluate_checkpoints(checkpoint_infos: List[dict],
                         num_games: Optional[int] = None,
                         league: Optional[League] = None,
                         num_workers: Optional[int] = None,
                         k_history: Optional[int] = None,
                         show_progress: bool = True):
    """Evaluate checkpoints head-to-head using the parallel eval pipeline."""
    results = {}
    matchups = list(itertools.combinations(checkpoint_infos, 2))
    total_pairs = len(matchups)

    env_default_games = int(os.environ.get('SCOPONE_EVAL_GAMES', '1000') or 1000)
    games_to_play = int(num_games) if (num_games is not None and num_games > 0) else env_default_games
    games_to_play = max(1, games_to_play)

    env_workers_default = int(os.environ.get('SCOPONE_EVAL_WORKERS', str(max(1, (os.cpu_count() or 1))))) or 1
    effective_workers = max(1, int(num_workers)) if num_workers is not None else max(1, env_workers_default)

    env_k_history = int(os.environ.get('SCOPONE_EVAL_K_HISTORY', '39') or 39)
    effective_k_history = int(k_history) if k_history is not None else env_k_history

    mcts_cfg = _build_eval_mcts_cfg()
    belief_particles = int(os.environ.get('SCOPONE_EVAL_BELIEF_PARTICLES', '0') or 0)
    belief_ess = float(os.environ.get('SCOPONE_EVAL_BELIEF_ESS_FRAC', '0.5') or 0.5)

    print(f"[benchmark] Using {effective_workers} worker(s) | games per matchup={games_to_play} | k_history={effective_k_history}")

    prev_workers_env_global = os.environ.get('SCOPONE_EVAL_WORKERS')
    os.environ['SCOPONE_EVAL_WORKERS'] = str(max(1, effective_workers))
    try:
        for pair_idx, (info_a, info_b) in enumerate(matchups, start=1):
            name_a = info_a["name"]
            name_b = info_b["name"]
            path_a = info_a["file_path"]
            path_b = info_b["file_path"]

            print(f"\nEvaluating ({pair_idx}/{total_pairs}): {name_a} vs {name_b}")
            tqdm_desc = f"{name_a} vs {name_b}"
            use_parallel = effective_workers > 1 and games_to_play > 1

            if use_parallel:
                diff_eval, breakdown = evaluate_pair_actors_parallel(
                    path_a,
                    path_b,
                    games=games_to_play,
                    k_history=effective_k_history,
                    mcts=mcts_cfg,
                    belief_particles=belief_particles,
                    belief_ess_frac=belief_ess,
                    num_workers=effective_workers,
                    tqdm_desc=tqdm_desc,
                    tqdm_disable=not show_progress,
                )
            else:
                prev_local_workers = os.environ.get('SCOPONE_EVAL_WORKERS')
                os.environ['SCOPONE_EVAL_WORKERS'] = '1'
                try:
                    diff_eval, breakdown = evaluate_pair_actors(
                        path_a,
                        path_b,
                        games=games_to_play,
                        k_history=effective_k_history,
                        mcts=mcts_cfg,
                        belief_particles=belief_particles,
                        belief_ess_frac=belief_ess,
                        tqdm_desc=tqdm_desc,
                        tqdm_position=0,
                        tqdm_disable=not show_progress,
                    )
                finally:
                    if prev_local_workers is None:
                        os.environ.pop('SCOPONE_EVAL_WORKERS', None)
                    else:
                        os.environ['SCOPONE_EVAL_WORKERS'] = prev_local_workers

            meta = breakdown.get('meta') or {}
            win_rate_a = max(0.0, min(1.0, float(meta.get('win_rate_agent1', meta.get('win_rate', 0.0)))))
            win_rate_b = max(0.0, min(1.0, float(meta.get('win_rate_agent2', 1.0 - win_rate_a))))
            draw_rate = max(0.0, min(1.0, float(meta.get('draw_rate', 1.0 - win_rate_a - win_rate_b))))
            if 'actor_avg_score_agent1' in meta and 'actor_avg_score_agent2' in meta:
                avg_score_a = float(meta.get('actor_avg_score_agent1', 0.0))
                avg_score_b = float(meta.get('actor_avg_score_agent2', 0.0))
            else:
                team_a_stats = breakdown.get(0, {}) or {}
                team_b_stats = breakdown.get(1, {}) or {}
                avg_score_a = float(team_a_stats.get('total', 0.0))
                avg_score_b = float(team_b_stats.get('total', 0.0))
            diff_from_scores = avg_score_a - avg_score_b

        matchup_key = f"{name_a}_vs_{name_b}"
        results[matchup_key] = {
            "games": games_to_play,
            "win_rate_agent1": win_rate_a,
            "win_rate_agent2": win_rate_b,
            "draw_rate": draw_rate,
            "diff_avg": float(diff_from_scores),
            "agent1_avg_score": avg_score_a,
            "agent2_avg_score": avg_score_b,
            "breakdown": breakdown,
            "agent1_path": info_a["league_key"],
            "agent2_path": info_b["league_key"],
            "agent1_name": name_a,
            "agent2_name": name_b,
        }

        print(f"  Win rate {name_a}: {win_rate_a * 100:.1f}%")
        print(f"  Win rate {name_b}: {win_rate_b * 100:.1f}%")
        print(f"  Draw rate: {draw_rate * 100:.1f}%")
        print(f"  Avg score {name_a}: {avg_score_a:.2f}")
        print(f"  Avg score {name_b}: {avg_score_b:.2f}")
        print(f"  Avg diff ({name_a} - {name_b}): {diff_from_scores:.2f}")

        if league is not None:
            league.update_elo_from_diff(info_a["league_key"], info_b["league_key"], diff_from_scores)
            try:
                league._save()
            except Exception:
                pass
            elo_a = league.elo.get(info_a["league_key"], 1000.0)
            elo_b = league.elo.get(info_b["league_key"], 1000.0)
        print(f"  Updated league Elo: {name_a} -> {elo_a:.1f}, {name_b} -> {elo_b:.1f}")

    finally:
        if prev_workers_env_global is None:
            os.environ.pop('SCOPONE_EVAL_WORKERS', None)
        else:
            os.environ['SCOPONE_EVAL_WORKERS'] = prev_workers_env_global

    return results




def generate_excel_comparison(checkpoint_infos: List[dict], results, output_file):
    """
    Generate an Excel file with comparative results between models.
    Each metric has its own dedicated sheet.
    """
    checkpoints_info = []
    for info in checkpoint_infos:
        file_path = info["file_path"]
        episode = extract_episode_number(file_path)
        name = info["name"]
        checkpoints_info.append((name, episode, file_path))

    checkpoints_info.sort(key=lambda x: x[1])
    model_names = [name for name, _, _ in checkpoints_info]

    def _empty_matrix():
        return pd.DataFrame(float('nan'), index=model_names, columns=model_names)

    win_rate_matrix = _empty_matrix()
    score_diff_matrix = _empty_matrix()
    avg_score_matrix = _empty_matrix()
    draw_rate_matrix = _empty_matrix()

    for data in results.values():
        name_a = data["agent1_name"]
        name_b = data["agent2_name"]
        win_rate_matrix.loc[name_a, name_b] = data["win_rate_agent1"] * 100.0
        win_rate_matrix.loc[name_b, name_a] = data["win_rate_agent2"] * 100.0
        score_diff_matrix.loc[name_a, name_b] = data["diff_avg"]
        score_diff_matrix.loc[name_b, name_a] = -data["diff_avg"]
        avg_score_matrix.loc[name_a, name_b] = data["agent1_avg_score"]
        avg_score_matrix.loc[name_b, name_a] = data["agent2_avg_score"]
        draw_rate_matrix.loc[name_a, name_b] = data["draw_rate"] * 100.0
        draw_rate_matrix.loc[name_b, name_a] = data["draw_rate"] * 100.0

    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Summary"

    summary_sheet['A1'] = "Checkpoint Benchmark Summary"
    summary_sheet['A1'].font = Font(bold=True, size=14)
    summary_sheet.merge_cells('A1:E1')

    headers = ["Model", "Episodes", "Average Win Rate (%)", "Average Score Diff", "Average Draw Rate (%)", "Path"]
    for col_idx, header in enumerate(headers, start=1):
        cell = summary_sheet.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)

    for row_idx, (name, episode, path_value) in enumerate(checkpoints_info, start=4):
        wins_row = win_rate_matrix.loc[name].dropna()
        diffs_row = score_diff_matrix.loc[name].dropna()
        draws_row = draw_rate_matrix.loc[name].dropna()
        avg_win_rate = wins_row.mean() if not wins_row.empty else float('nan')
        avg_score_diff = diffs_row.mean() if not diffs_row.empty else float('nan')
        avg_draw_rate = draws_row.mean() if not draws_row.empty else float('nan')

        summary_sheet.cell(row=row_idx, column=1, value=name)
        summary_sheet.cell(row=row_idx, column=2, value=episode if episode != float('inf') else "Unknown")
        cell_win = summary_sheet.cell(row=row_idx, column=3,
                                      value=None if pd.isna(avg_win_rate) else round(avg_win_rate, 1))
        cell_diff = summary_sheet.cell(row=row_idx, column=4,
                                       value=None if pd.isna(avg_score_diff) else round(avg_score_diff, 2))
        cell_draw = summary_sheet.cell(row=row_idx, column=5,
                                       value=None if pd.isna(avg_draw_rate) else round(avg_draw_rate, 1))
        summary_sheet.cell(row=row_idx, column=6, value=path_value)
        if cell_win.value is not None:
            cell_win.number_format = "0.0"
        if cell_diff.value is not None:
            cell_diff.number_format = "0.00"
        if cell_draw.value is not None:
            cell_draw.number_format = "0.0"

    for col in range(1, len(headers) + 1):
        column_letter = openpyxl.utils.get_column_letter(col)
        max_length = 0
        for row in range(1, summary_sheet.max_row + 1):
            value = summary_sheet.cell(row=row, column=col).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        summary_sheet.column_dimensions[column_letter].width = max_length + 2

    metrics = [
        ("Win Rates", win_rate_matrix, "Win Rates (Row vs Column)",
         "Percentage of games won by the row model against the column model", "{:.1f}%", True),
        ("Draw Rates", draw_rate_matrix, "Draw Rates (Row vs Column)",
         "Percentage of games ending in a draw for each matchup", "{:.1f}%", False),
        ("Score Differences", score_diff_matrix, "Score Differences (Row minus Column)",
         "Positive values mean the row model scores higher than the column model", "{:.2f}", True),
        ("Average Scores", avg_score_matrix, "Average Total Score (Row perspective)",
         "Average total score achieved by the row model in each matchup", "{:.2f}", False),
    ]

    for sheet_name, matrix, title, description, fmt, apply_gradient in metrics:
        sheet = wb.create_sheet(title=sheet_name)
        sheet['A1'] = title
        sheet['A1'].font = Font(bold=True, size=14)
        last_col_letter = openpyxl.utils.get_column_letter(len(model_names) + 1)
        sheet.merge_cells(f'A1:{last_col_letter}1')
        sheet['A2'] = description
        sheet.merge_cells(f'A2:{last_col_letter}2')

        sheet.cell(row=3, column=1, value="Model").font = Font(bold=True)
        for col_idx, model in enumerate(model_names, start=2):
            sheet.cell(row=3, column=col_idx, value=model).font = Font(bold=True)

        for row_idx, row_model in enumerate(model_names, start=4):
            sheet.cell(row=row_idx, column=1, value=row_model).font = Font(bold=True)
            for col_idx, col_model in enumerate(model_names, start=2):
                numeric_val = matrix.loc[row_model, col_model]
                display = "-" if pd.isna(numeric_val) else fmt.format(numeric_val)
                cell = sheet.cell(row=row_idx, column=col_idx, value=display)
                if apply_gradient and not pd.isna(numeric_val):
                    value_f = float(numeric_val)
                    if sheet_name == "Win Rates":
                        midpoint = 50.0
                        if value_f >= midpoint:
                            intensity = min(255, int(155 + (value_f - midpoint) * 2))
                            green_hex = format(intensity, '02x')
                            red_hex = format(255 - intensity // 3, '02x')
                            cell.fill = PatternFill(start_color=f"{red_hex}{green_hex}55",
                                                    end_color=f"{red_hex}{green_hex}55",
                                                    fill_type="solid")
                        else:
                            intensity = min(255, int(155 + (midpoint - value_f) * 2))
                            red_hex = format(intensity, '02x')
                            green_hex = format(255 - intensity // 3, '02x')
                            cell.fill = PatternFill(start_color=f"{red_hex}{green_hex}55",
                                                    end_color=f"{red_hex}{green_hex}55",
                                                    fill_type="solid")
                    elif sheet_name == "Score Differences":
                        if value_f > 0:
                            intensity = min(255, int(155 + min(value_f * 25, 100)))
                            green_hex = format(intensity, '02x')
                            red_hex = format(255 - intensity // 3, '02x')
                            cell.fill = PatternFill(start_color=f"{red_hex}{green_hex}55",
                                                    end_color=f"{red_hex}{green_hex}55",
                                                    fill_type="solid")
                        elif value_f < 0:
                            intensity = min(255, int(155 + min(abs(value_f) * 25, 100)))
                            red_hex = format(intensity, '02x')
                            green_hex = format(255 - intensity // 3, '02x')
                            cell.fill = PatternFill(start_color=f"{red_hex}{green_hex}55",
                                                    end_color=f"{red_hex}{green_hex}55",
                                                    fill_type="solid")

        max_cols = len(model_names) + 1
        for col in range(1, max_cols + 1):
            column_letter = openpyxl.utils.get_column_letter(col)
            max_length = 0
            for row in range(1, sheet.max_row + 1):
                value = sheet.cell(row=row, column=col).value
                if value is not None:
                    max_length = max(max_length, len(str(value)))
            sheet.column_dimensions[column_letter].width = max_length + 2

    detailed_sheet = wb.create_sheet(title="Detailed Results")
    detailed_sheet['A1'] = "Detailed Matchup Results"
    detailed_sheet['A1'].font = Font(bold=True, size=14)
    detailed_sheet.merge_cells('A1:H1')

    detailed_headers = [
        "Model A", "Model B", "Games", "Win Rate A (%)",
        "Win Rate B (%)", "Draw Rate (%)", "Avg Score A", "Avg Score B", "Avg Diff (A-B)"
    ]
    for col_idx, header in enumerate(detailed_headers, start=1):
        detailed_sheet.cell(row=3, column=col_idx, value=header).font = Font(bold=True)

    row_idx = 4
    for data in results.values():
        detailed_sheet.cell(row=row_idx, column=1, value=data['agent1_name'])
        detailed_sheet.cell(row=row_idx, column=2, value=data['agent2_name'])
        detailed_sheet.cell(row=row_idx, column=3, value=data['games'])
        cell_wra = detailed_sheet.cell(row=row_idx, column=4, value=data['win_rate_agent1'] * 100.0)
        cell_wra.number_format = "0.0"
        cell_wrb = detailed_sheet.cell(row=row_idx, column=5, value=data['win_rate_agent2'] * 100.0)
        cell_wrb.number_format = "0.0"
        cell_draw = detailed_sheet.cell(row=row_idx, column=6, value=data['draw_rate'] * 100.0)
        cell_draw.number_format = "0.0"
        cell_sca = detailed_sheet.cell(row=row_idx, column=7, value=data['agent1_avg_score'])
        cell_sca.number_format = "0.00"
        cell_scb = detailed_sheet.cell(row=row_idx, column=8, value=data['agent2_avg_score'])
        cell_scb.number_format = "0.00"
        cell_diff = detailed_sheet.cell(row=row_idx, column=9, value=data['diff_avg'])
        cell_diff.number_format = "0.00"
        row_idx += 1

    for col in range(1, len(detailed_headers) + 1):
        column_letter = openpyxl.utils.get_column_letter(col)
        max_length = 0
        for row in range(1, detailed_sheet.max_row + 1):
            value = detailed_sheet.cell(row=row, column=col).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        detailed_sheet.column_dimensions[column_letter].width = max_length + 2

    wb.save(output_file)
    print(f"\nEnhanced Excel comparison report saved to {output_file}")

    return output_file

def main():
    parser = argparse.ArgumentParser(description="Benchmark Scopone checkpoints head-to-head and update league Elo")
    parser.add_argument("--checkpoints", nargs="+", help="Paths to checkpoint files or directories")
    parser.add_argument("--checkpoint_dir", help="Directory containing checkpoints")
    parser.add_argument("--checkpoint_pattern", default="*.pth", help="Pattern to match checkpoint files inside directories")
    parser.add_argument("--games", type=int, default=1000, help="Number of games to play for each matchup")
    parser.add_argument("--output", help="Output file path (default: auto-generated)")
    parser.add_argument("--excel", help="Excel output file path (default: auto-generated)")
    parser.add_argument("--limit", type=int, help="Limit the number of checkpoints to evaluate")
    parser.add_argument("--workers", type=int, help="Number of parallel evaluation workers")
    parser.add_argument("--k-history", type=int, help="Observation history length for evaluation")
    parser.add_argument("--no-progress", action='store_true', help="Disable per-match evaluation progress bars")
    parser.add_argument("--show-progress", action='store_true', help=argparse.SUPPRESS)
    args = parser.parse_args()

    checkpoint_paths: List[str] = []

    if args.checkpoints:
        for cp in args.checkpoints:
            if os.path.isfile(cp):
                checkpoint_paths.append(cp)
            elif os.path.isdir(cp):
                checkpoint_paths.extend(find_checkpoints(cp, args.checkpoint_pattern))
            else:
                print(f"Warning: Checkpoint not found: {cp}")

    if args.checkpoint_dir:
        checkpoint_paths.extend(find_checkpoints(args.checkpoint_dir, args.checkpoint_pattern))

    default_dir = os.path.join('checkpoints')
    if os.path.isdir(default_dir):
        checkpoint_paths.extend(find_checkpoints(default_dir, args.checkpoint_pattern))

    normalized_paths: List[str] = []
    seen_paths = set()
    for cp in checkpoint_paths:
        norm_cp = os.path.normpath(cp)
        if norm_cp not in seen_paths:
            seen_paths.add(norm_cp)
            normalized_paths.append(norm_cp)
    checkpoint_paths = normalized_paths

    existing_files: List[str] = []
    for cp in checkpoint_paths:
        if os.path.isfile(cp):
            existing_files.append(cp)
        else:
            print(f"Warning: Skipping missing checkpoint: {cp}")
    checkpoint_paths = existing_files

    if args.limit and len(checkpoint_paths) > args.limit:
        step = max(1, len(checkpoint_paths) // args.limit)
        checkpoint_paths = [checkpoint_paths[i] for i in range(0, len(checkpoint_paths), step)][:args.limit]

    if not checkpoint_paths:
        print("Error: No checkpoint files found. Please provide valid checkpoint paths.")
        return

    league = League()
    canonicalize_league_paths(league)

    raw_infos: List[dict] = []
    for cp in checkpoint_paths:
        league_key = os.path.normpath(cp)
        if league_key not in getattr(league, 'history', []):
            league.register(league_key)
        raw_infos.append({
            "file_path": cp,
            "league_key": league_key,
            "base_name": checkpoint_display_name(cp),
        })

    name_counts = {}
    checkpoint_infos: List[dict] = []
    for info in raw_infos:
        base = info["base_name"]
        count = name_counts.get(base, 0)
        display_name = base if count == 0 else f"{base}_{count + 1}"
        name_counts[base] = count + 1
        checkpoint_infos.append({
            "name": display_name,
            "file_path": info["file_path"],
            "league_key": info["league_key"],
            "base_name": base,
        })

    print(f"Found {len(checkpoint_infos)} checkpoint files:")
    for idx, info in enumerate(checkpoint_infos, start=1):
        print(f"{idx}. {info['file_path']} [{info['name']}]")

    workers_env = os.environ.get('SCOPONE_EVAL_WORKERS')
    if args.workers is not None:
        num_workers = max(1, args.workers)
    elif workers_env:
        try:
            num_workers = max(1, int(workers_env))
        except ValueError:
            num_workers = 1
    else:
        cpu_count = os.cpu_count() or 1
        num_workers = max(1, cpu_count // 2)
    num_workers = max(1, num_workers)

    k_history_env = os.environ.get('SCOPONE_EVAL_K_HISTORY') or os.environ.get('SCOPONE_EVAL_KH')
    if args.k_history is not None:
        k_history = max(1, args.k_history)
    elif k_history_env:
        try:
            k_history = max(1, int(k_history_env))
        except ValueError:
            k_history = 39
    else:
        k_history = 39

    progress_enabled = True
    if getattr(args, 'no_progress', False):
        progress_enabled = False
    elif getattr(args, 'show_progress', False):
        progress_enabled = True

    if torch.cuda.is_available():
        torch.backends.cudnn.benchmark = True
        torch.backends.cudnn.enabled = True
        if hasattr(torch.backends.cuda, 'matmul'):
            torch.backends.cuda.matmul.allow_tf32 = True
        torch.backends.cudnn.allow_tf32 = True
        torch.cuda.empty_cache()

    timestamp = time.strftime("%Y%m%d-%H%M%S")
    output_file = args.output if args.output else f"benchmark_comparison_{timestamp}.txt"
    excel_file = args.excel if args.excel else f"benchmark_comparison_{timestamp}.xlsx"
    excel_generated = False

    with open(output_file, "w") as f:
        f.write(f"Scopone AI Checkpoint Benchmark Results ({time.strftime('%Y-%m-%d %H:%M:%S')})\n")
        f.write(f"Games per matchup: {args.games}\n")
        f.write(f"Eval k_history: {k_history}\n")
        f.write(f"Parallel workers: {num_workers}\n\n")
        f.write(f"Checkpoints evaluated ({len(checkpoint_infos)}):\n")
        for idx, info in enumerate(checkpoint_infos, start=1):
            f.write(f"{idx}. {info['file_path']} [{info['name']}]\n")
        f.write("\n")

        comparison_results = {}
        if len(checkpoint_infos) > 1:
            f.write("=== COMPARISON RESULTS ===\n\n")
            comparison_results = evaluate_checkpoints(
                checkpoint_infos,
                args.games,
                league=league,
                num_workers=num_workers,
                k_history=k_history,
                show_progress=progress_enabled,
            )

            for data in comparison_results.values():
                name_a = data['agent1_name']
                name_b = data['agent2_name']
                win_rate_a_pct = data['win_rate_agent1'] * 100.0
                win_rate_b_pct = data['win_rate_agent2'] * 100.0
                diff_avg = data['diff_avg']
                elo_a = league.elo.get(data['agent1_path'], 1000.0)
                elo_b = league.elo.get(data['agent2_path'], 1000.0)

                draw_pct = data['draw_rate'] * 100.0

                f.write(f"{name_a} vs {name_b}:\n")
                f.write(f"  Win rate {name_a}: {win_rate_a_pct:.1f}%\n")
                f.write(f"  Win rate {name_b}: {win_rate_b_pct:.1f}%\n")
                f.write(f"  Draw rate: {draw_pct:.1f}%\n")
                f.write(f"  Avg score {name_a}: {data['agent1_avg_score']:.2f}\n")
                f.write(f"  Avg score {name_b}: {data['agent2_avg_score']:.2f}\n")
                f.write(f"  Avg diff ({name_a} - {name_b}): {diff_avg:.2f}\n")
                f.write(f"  Updated league Elo: {name_a} -> {elo_a:.1f}, {name_b} -> {elo_b:.1f}\n\n")

            generate_excel_comparison(checkpoint_infos, comparison_results, excel_file)
            excel_generated = True

        f.write("=== LEAGUE ELO SNAPSHOT ===\n")
        for info in checkpoint_infos:
            elo_value = league.elo.get(info["league_key"], 1000.0)
            f.write(f"{info['name']}: {elo_value:.1f} ({info['file_path']})\n")

    print(f"\nResults saved to {output_file}")
    if excel_generated:
        print(f"Excel comparison report saved to {excel_file}")
    elif len(checkpoint_infos) > 1:
        print("Excel comparison report was not generated.")


if __name__ == "__main__":
    main()
